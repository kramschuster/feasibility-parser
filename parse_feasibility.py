#!/usr/bin/env python3
"""
Parse Sagimet SB2640 feasibility questionnaire(s) and add rows to the tracker Excel.

Usage – single file:
    python3 parse_feasibility.py "Miller Feasibility.docx"
    python3 parse_feasibility.py "Miller Feasibility.docx" --dry-run

Usage – whole directory:
    python3 parse_feasibility.py /path/to/docs/
    python3 parse_feasibility.py /path/to/docs/ --dry-run

The script looks up the investigator name / site address from the master spreadsheet
using the last name in the docx file name (e.g. "Miller" from "Miller Feasibility.docx").
Skips any .docx files whose last name is not found in the master spreadsheet, and prints
a summary of successes and failures at the end.
"""

import sys
import os
import re
import argparse
import shutil
from datetime import datetime
from docx import Document
from docx.oxml.ns import qn
import openpyxl
from copy import copy

if getattr(sys, "frozen", False):
    # Running as a PyInstaller .exe — Excel files live next to the executable
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TRACKER_FILE = os.path.join(BASE_DIR, "Sagimet SB2640 Feasibility Tracker.xlsx")
MASTER_FILE  = os.path.join(BASE_DIR, "INTERNAL Acne Phase 3 Feasibility - Master US - email-previous study info.xlsx")

# Texts that are empty placeholders in the form
PLACEHOLDERS = {
    "please enter comment here.",
    "please enter text here.",
    "please enter information here.",
    "please enter integer number here.",
    "comment",
    "please explain.",
    "please comment how patient records would be provided for monitoring",
    "please comment how patient records would be provided for monitoring.",
    "n/a",
    "please select all that apply",
}

def is_placeholder(text):
    t = text.strip().lower()
    return not t or t in PLACEHOLDERS or (t.startswith("please enter") and len(t) < 70)

# ── Document walking ─────────────────────────────────────────────────────────

def get_checked(element):
    """Return True/False if element has a checkbox, None otherwise."""
    xml = getattr(element, "xml", "")
    m = re.findall(r'<w14:checked w14:val="(\d+)"', xml)
    return (m[0] == "1") if m else None

def walk_body(element, items=None):
    """Collect all paragraphs in body order (including inside top-level SDTs)."""
    if items is None:
        items = []
    tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag

    if tag == "body":
        for child in element:
            walk_body(child, items)
    elif tag == "p":
        text = "".join(t.text or "" for t in element.iter(qn("w:t"))).strip()
        chk  = get_checked(element)
        items.append({"text": text, "checked": chk})
    elif tag == "sdt":
        content = element.find(qn("w:sdtContent"))
        if content is not None:
            has_p = any(c.tag.split("}")[-1] == "p" for c in content)
            if has_p:
                for child in content:
                    walk_body(child, items)
            # else: inline SDT – its text is already included in the parent paragraph
    # tbl: skip (only header tables in this doc)
    return items

# ── Generic helpers ──────────────────────────────────────────────────────────

def _normalize(text):
    """Normalize curly quotes/apostrophes to straight for comparison."""
    return text.replace("’", "'").replace("‘", "'").replace("“", '"').replace("”", '"')

def find_idx(items, fragment, start=0):
    frag = _normalize(fragment).lower()
    for i in range(start, len(items)):
        if frag in _normalize(items[i]["text"]).lower():
            return i
    return -1

def get_yes_no_after(items, idx):
    """Return 'Yes'/'No' by checking YES/NO checkboxes after idx.
    Requires at least one checkbox to actually be checked; returns None
    if both checkboxes are found unchecked (i.e. question left blank).
    """
    yes_checked = no_checked = None
    for j in range(idx + 1, min(idx + 6, len(items))):
        item = items[j]
        if item["checked"] is None:
            continue
        txt = item["text"].upper()
        if "YES" in txt and yes_checked is None:
            yes_checked = item["checked"]
        elif "NO" in txt and "NOT SURE" not in txt and no_checked is None:
            no_checked = item["checked"]
        if yes_checked is not None and no_checked is not None:
            break
    if yes_checked:
        return "Yes"
    if no_checked:
        return "No"
    # Neither checked – return None so caller can decide
    return None

def get_not_sure_after(items, idx):
    """Return 'Not sure' / 'Yes' / 'No' for a 3-way question."""
    answers = {}
    for j in range(idx + 1, min(idx + 8, len(items))):
        item = items[j]
        if item["checked"] is None:
            break
        t = item["text"].upper()
        if "YES" in t:
            answers["yes"] = item["checked"]
        elif "NOT SURE" in t:
            answers["not_sure"] = item["checked"]
        elif "NO" in t:
            answers["no"] = item["checked"]
    if answers.get("yes"):
        return "Yes"
    if answers.get("not_sure"):
        return "Not sure"
    if answers.get("no"):
        return "No"
    return None

def get_comment_after_yn(items, yn_idx):
    """Return comment text immediately after YES/NO checkboxes, or None."""
    # skip YES + NO paragraphs (they are checkboxes)
    j = yn_idx + 1
    while j < len(items) and items[j]["checked"] is not None:
        j += 1
    if j < len(items):
        txt = items[j]["text"]
        if not is_placeholder(txt) and items[j]["checked"] is None:
            return txt
    return None

def get_next_real_text(items, start, max_lines=1, stop_frags=None):
    """Return the next non-placeholder paragraph(s) after start, joining with \\n."""
    stop_frags = [_normalize(f).lower() for f in (stop_frags or [])]
    collected = []
    for j in range(start + 1, len(items)):
        item = items[j]
        if item["checked"] is not None:
            break  # hit a checkbox → stop
        txt = item["text"]
        ntxt = _normalize(txt).lower()
        if any(sf in ntxt for sf in stop_frags):
            break
        if not is_placeholder(txt):
            collected.append(txt)
            if len(collected) >= max_lines:
                break
    return "\n".join(collected) if collected else None

def get_inline_after(text, sep):
    """Extract text after `sep` in `text`, or None."""
    if sep.lower() in text.lower():
        idx = text.lower().index(sep.lower())
        return text[idx + len(sep):].strip() or None
    return None

def get_multi_checkbox(items, question_idx, stop_frags=None):
    """Return comma-joined labels of checked checkboxes after question_idx."""
    stop_frags = [f.lower() for f in (stop_frags or [])]
    checked_labels = []
    for j in range(question_idx + 1, len(items)):
        item = items[j]
        if item["checked"] is None:
            if stop_frags and any(sf in item["text"].lower() for sf in stop_frags):
                break
            if not is_placeholder(item["text"]) and item["text"]:
                break  # non-checkbox, non-placeholder → next question
            continue
        if item["checked"]:
            label = item["text"].replace("☒", "").replace("☐", "").strip()
            # strip trailing whitespace / tab artifacts
            label = re.sub(r"\s+", " ", label).rstrip()
            if label and not is_placeholder(label):
                checked_labels.append(label)
    return ", ".join(checked_labels) if checked_labels else None

# ── Master spreadsheet lookup ────────────────────────────────────────────────

def find_master_info(last_name):
    wb = openpyxl.load_workbook(MASTER_FILE)
    ws = wb["Internal -Acne"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        inv = row[5]
        if inv and last_name.lower() in str(inv).lower():
            return {
                "investigator": str(inv).strip(),
                "site_address": str(row[6]).strip() if row[6] else "",
                "state":        str(row[7]).strip() if row[7] else "",
            }
    return None

# ── Main parser ──────────────────────────────────────────────────────────────

def parse_docx(docx_path):
    doc   = Document(docx_path)
    items = walk_body(doc.element.body)

    # Find questionnaire start (Q1)
    q1_idx = find_idx(items, "Based on your review of the current study design")
    if q1_idx == -1:
        raise ValueError("Could not find Q1 in document. Is this the right file?")
    items = items[q1_idx:]  # work only in questionnaire section

    def fi(frag, start=0):
        return find_idx(items, frag, start)

    data = [None] * 78  # 78 Excel columns (0–77)

    # ── Q1: feasible ────────────────────────────────────────────────────────
    q1i = fi("Based on your review of the current study design")
    data[3] = get_yes_no_after(items, q1i)

    # ── Q2: interested ──────────────────────────────────────────────────────
    q2i = fi("Would your site be interested in participating in this study")
    data[4] = get_yes_no_after(items, q2i)

    # ── Q2a: decline reason ─────────────────────────────────────────────────
    q2a_i = fi("If your site is declining this study, please provide a brief reason why")
    if q2a_i != -1:
        # answer is the next non-placeholder paragraph (or inline in the same para)
        txt = items[q2a_i]["text"]
        inline = get_inline_after(txt, "why.")
        if inline and not is_placeholder(inline):
            data[5] = inline
        else:
            # next item may be a checkbox with comment text
            j = q2a_i + 1
            if j < len(items) and items[j]["checked"] is not None:
                t = items[j]["text"].replace("☐", "").replace("☒", "").strip()
                if not is_placeholder(t):
                    data[5] = t

    # ── Q3: board-certified ─────────────────────────────────────────────────
    q3i = fi("Is the PI a board-certified Dermatologist")
    data[6] = get_yes_no_after(items, q3i)

    # ── Q4: institution contact info ─────────────────────────────────────────
    q4i = fi("Please provide institution contact information")
    if q4i != -1:
        fields = {}
        for j in range(q4i + 1, min(q4i + 20, len(items))):
            txt = items[j]["text"]
            for label in ["Company", "Address 1", "Address 2", "City/Town",
                          "State/Province", "ZIP/Postal Code", "Country",
                          "Email Address", "Phone Number", "Fax Number"]:
                if txt.startswith(label):
                    val = txt[len(label):].strip()
                    if val and not is_placeholder(val):
                        fields[label] = val
                    break
            # Stop when we hit Q5
            if "best describes your site" in txt.lower():
                break

        parts = []
        for label in ["Company", "Address 1", "Address 2", "City/Town",
                       "State/Province", "ZIP/Postal Code", "Country"]:
            if fields.get(label):
                parts.append(fields[label])
        addr_block = "\n".join(parts)

        # Append email / phone / fax
        contact_parts = []
        for label, key in [("Email Address", "Email Address"),
                            ("Phone Number", "Phone Number"),
                            ("Fax Number",   "Fax Number")]:
            if fields.get(label):
                contact_parts.append(f"{label}: {fields[label]}")
        if contact_parts:
            addr_block += ("\n" if addr_block else "") + "\n".join(contact_parts)
        data[7] = addr_block or None

    # ── Q5: site setting ─────────────────────────────────────────────────────
    q5i = fi("best describes your site")
    if q5i != -1:
        opts = []
        for j in range(q5i + 1, min(q5i + 10, len(items))):
            item = items[j]
            if item["checked"] is None:
                break
            if item["checked"]:
                label = re.sub(r"[☒☐]", "", item["text"]).strip()
                # Keep only the first word-group before any embedded comment
                # e.g. "SMO  Site has activities as International SMO"
                label = label.split("  ")[0].strip()
                if label:
                    opts.append(label)
        data[8] = ", ".join(opts) if opts else None

    # ── Q6: PI contact ──────────────────────────────────────────────────────
    q6i = fi("Please enter the Principal Investigator")
    if q6i != -1:
        fields = {}
        for j in range(q6i + 1, min(q6i + 15, len(items))):
            txt = items[j]["text"]
            for label in ["Title", "First Name", "Last Name",
                          "Email Address", "Phone Number", "Fax Number", "Medical Specialty"]:
                if txt.startswith(label):
                    val = txt[len(label):].strip()
                    if val and not is_placeholder(val):
                        fields[label] = val
                    break
            if "all functions and number of study staff" in txt.lower():
                break

        lines = []
        for label in ["Title", "First Name", "Last Name", "Email Address",
                       "Phone Number", "Fax Number", "Medical Specialty"]:
            if fields.get(label):
                lines.append(f"{label}: {fields[label]}")
        data[9] = "\n".join(lines) or None

    # ── Q7: study staff ──────────────────────────────────────────────────────
    q7i = fi("all functions and number of study staff")
    if q7i != -1:
        # Collect until "How many years"
        staff_lines = []
        for j in range(q7i + 1, len(items)):
            if "how many years has the principal investigator" in items[j]["text"].lower():
                break
            if items[j]["checked"] is not None:
                break
            txt = items[j]["text"]
            if txt and not is_placeholder(txt):
                staff_lines.append(txt)
        data[10] = "\n".join(staff_lines) or None

    # ── Q8: years conducting trials ──────────────────────────────────────────
    q8i = fi("How many years has the Principal Investigator been conducting")
    if q8i != -1:
        data[11] = get_next_real_text(items, q8i)

    # ── Q9: acne studies ─────────────────────────────────────────────────────
    q9i = fi("Has the PI participated in Ph II and/or Ph III Acne Vulgaris studies before")
    data[12] = get_yes_no_after(items, q9i)

    # Q9a: Ph II count (inline)
    q9a_i = fi("If yes, how many Ph II studies?")
    if q9a_i != -1:
        txt = items[q9a_i]["text"]
        val = get_inline_after(txt, "Ph II studies?")
        if val and not is_placeholder(val):
            data[13] = val

    # Q9b: Ph III count (inline)
    q9b_i = fi("If yes, how many Ph III studies?")
    if q9b_i != -1:
        txt = items[q9b_i]["text"]
        val = get_inline_after(txt, "Ph III studies?")
        if val and not is_placeholder(val):
            data[14] = val

    # ── Q10: GCP training ────────────────────────────────────────────────────
    q10i = fi("Has the Principal Investigator been trained in GCP")
    data[15] = get_yes_no_after(items, q10i)

    # Q10 date (inline)
    q10d = fi("Please provide date of last training:")
    if q10d != -1:
        txt = items[q10d]["text"]
        val = get_inline_after(txt, "last training:")
        if val:
            # Strip "MM/YY" placeholder prefix if present
            val = re.sub(r"^MM/YY[:\s]*", "", val, flags=re.IGNORECASE).strip()
        if val and not is_placeholder(val):
            data[16] = val

    # ── Q11: another investigator ────────────────────────────────────────────
    q11i = fi("Will another Investigator equal in qualification")
    data[17] = get_yes_no_after(items, q11i)

    # ── Q12: FDA audit ───────────────────────────────────────────────────────
    q12i = fi("Has your site been audited by the FDA in the past 5 years")
    data[18] = get_yes_no_after(items, q12i)

    # Q12a: received 483
    q12a_i = fi("Did your site receive a 483")
    data[19] = get_yes_no_after(items, q12a_i)

    # ── Q13: electronic patient files ────────────────────────────────────────
    q13i = fi("Does your site use electronic patient files")
    q13_ans = get_yes_no_after(items, q13i)
    data[20] = q13_ans

    # Q13a: which system – only relevant if Q13 = Yes
    q13a_i = fi("If yes, which electronic patient file system do you use")
    if q13a_i != -1 and q13_ans == "Yes":
        val = get_next_real_text(items, q13a_i,
                                  stop_frags=["If no, please comment on current process",
                                              "Would it be possible to grant"])
        if val and "if no" not in val.lower():
            data[21] = val

    # Q13b: paper process – only relevant if Q13 = No
    q13b_i = fi("If no, please comment on current process")
    if q13b_i != -1 and q13_ans != "Yes":
        txt = items[q13b_i]["text"]
        val = None
        for sep in ["current process.", "current process"]:
            raw = get_inline_after(txt, sep)
            if raw:
                raw = raw.lstrip(".,").strip()
                # Strip trailing placeholder text appended by SDT concatenation
                raw = re.sub(r'Please enter\b.*$', '', raw, flags=re.IGNORECASE).strip().rstrip(".,").strip()
                if raw and not is_placeholder(raw):
                    val = raw
                break
        if val:
            data[22] = val
        else:
            # Try next paragraph
            val = get_next_real_text(items, q13b_i,
                                      stop_frags=["Would it be possible to grant"])
            if val:
                val = re.sub(r'Please enter\b.*$', '', val, flags=re.IGNORECASE).strip().rstrip(".,").strip()
                if val and not is_placeholder(val):
                    data[22] = val

    # ── Q14: CRA access ─────────────────────────────────────────────────────
    q14i = fi("Would it be possible to grant the Clinical Research Associate access")
    if q14i != -1:
        yn_ans = get_yes_no_after(items, q14i)
        yn_idx = q14i
        for j in range(q14i + 1, min(q14i + 5, len(items))):
            if items[j]["checked"] is not None and "YES" in items[j]["text"].upper():
                yn_idx = j
                break
        comment = get_comment_after_yn(items, yn_idx)
        if yn_ans and comment:
            data[23] = f"{yn_ans}. {comment}"
        elif yn_ans:
            data[23] = yn_ans
        elif comment:
            # Both checkboxes unchecked but there is a comment (e.g. "printed copies...")
            data[23] = comment

    # ── Q15-Q19: simple YES/NO ───────────────────────────────────────────────
    q15i = fi("Does your site have a lockable, refrigerated and temperature-controlled area")
    data[24] = get_yes_no_after(items, q15i)

    q16i = fi("Is your site equipped with a freezer")
    data[25] = get_yes_no_after(items, q16i)

    # Q17: temp control + comment
    q17i = fi("Does your site have appropriate temperature-controlled storage and monitoring procedures")
    if q17i != -1:
        yn_ans = get_yes_no_after(items, q17i)
        yn_idx = q17i
        for j in range(q17i + 1, min(q17i + 5, len(items))):
            if items[j]["checked"] is not None and "YES" in items[j]["text"].upper():
                yn_idx = j
                break
        comment = get_comment_after_yn(items, yn_idx)
        if yn_ans and comment:
            data[26] = f"{yn_ans}. {comment}"
        elif yn_ans:
            data[26] = yn_ans
        elif comment:
            data[26] = comment

    q18i = fi("Is your site equipped with a refrigerator/freezer for blood sample storage")
    if q18i != -1:
        yn = get_yes_no_after(items, q18i)
        yn_idx = q18i
        for j in range(q18i + 1, min(q18i + 5, len(items))):
            if items[j]["checked"] is not None and "YES" in items[j]["text"].upper():
                yn_idx = j; break
        comment = get_comment_after_yn(items, yn_idx)
        data[27] = (f"{yn}. {comment}" if yn and comment else yn or comment)

    q19i = fi("Is your site familiar with sending blood samples to a central lab")
    if q19i != -1:
        yn = get_yes_no_after(items, q19i)
        yn_idx = q19i
        for j in range(q19i + 1, min(q19i + 5, len(items))):
            if items[j]["checked"] is not None and "YES" in items[j]["text"].upper():
                yn_idx = j; break
        comment = get_comment_after_yn(items, yn_idx)
        data[28] = (f"{yn}. {comment}" if yn and comment else yn or comment)

    # Q20: acne photography + comment
    q20i = fi("Does your site have experience with performing standardized central acne photography")
    if q20i != -1:
        yn = get_yes_no_after(items, q20i)
        yn_idx = q20i
        for j in range(q20i + 1, min(q20i + 5, len(items))):
            if items[j]["checked"] is not None and "YES" in items[j]["text"].upper():
                yn_idx = j; break
        comment = get_comment_after_yn(items, yn_idx)
        data[29] = (f"{yn}. {comment}" if yn and comment else yn or comment)

    # Q20a: house equipment + comment
    q20a_i = fi("Is your site able to house study required photography equipment")
    if q20a_i != -1:
        yn = get_yes_no_after(items, q20a_i)
        yn_idx = q20a_i
        for j in range(q20a_i + 1, min(q20a_i + 5, len(items))):
            if items[j]["checked"] is not None and "YES" in items[j]["text"].upper():
                yn_idx = j; break
        comment = get_comment_after_yn(items, yn_idx)
        data[30] = (f"{yn}. {comment}" if yn and comment else yn or comment)

    # Q21: sebutape
    q21i = fi("Does your site have experience performing Sebutape")
    if q21i != -1:
        yn = get_yes_no_after(items, q21i)
        yn_idx = q21i
        for j in range(q21i + 1, min(q21i + 5, len(items))):
            if items[j]["checked"] is not None and "YES" in items[j]["text"].upper():
                yn_idx = j; break
        comment = get_comment_after_yn(items, yn_idx)
        data[31] = (f"{yn}. {comment}" if yn and comment else yn or comment)

    # Q22: sebumeter experience
    q22i = fi("Does your site have experience using a Sebumeter")
    if q22i != -1:
        yn = get_yes_no_after(items, q22i)
        yn_idx = q22i
        for j in range(q22i + 1, min(q22i + 5, len(items))):
            if items[j]["checked"] is not None and "YES" in items[j]["text"].upper():
                yn_idx = j; break
        comment = get_comment_after_yn(items, yn_idx)
        data[32] = (f"{yn}. {comment}" if yn and comment else yn or comment)

    # Q23: calibrated sebumeter on-site + comment
    q23i = fi("Does your site have a calibrated Sebumeter")
    if q23i != -1:
        yn = get_yes_no_after(items, q23i)
        yn_idx = q23i
        for j in range(q23i + 1, min(q23i + 5, len(items))):
            if items[j]["checked"] is not None and "YES" in items[j]["text"].upper():
                yn_idx = j; break
        comment = get_comment_after_yn(items, yn_idx)
        data[33] = (f"{yn}. {comment}" if yn and comment else yn or comment)

    # Q23a: would require sebumeter
    q23a_i = fi("If no, would your site require a Sebumeter to be provided")
    data[34] = get_yes_no_after(items, q23a_i)

    # Q24: lesion count familiarity + comment
    q24i = fi("Is your site familiar with performing lesion count, IGA")
    if q24i != -1:
        yn = get_yes_no_after(items, q24i)
        yn_idx = q24i
        for j in range(q24i + 1, min(q24i + 5, len(items))):
            if items[j]["checked"] is not None and "YES" in items[j]["text"].upper():
                yn_idx = j; break
        comment = get_comment_after_yn(items, yn_idx)
        data[35] = (f"{yn}. {comment}" if yn and comment else yn or comment)

    # Q24a: who performs lesion count (inline after "?")
    q24a_i = fi("Who at the site will be performing the lesion count")
    if q24a_i != -1:
        txt = items[q24a_i]["text"]
        val = get_inline_after(txt, "lesion count?")
        if val and not is_placeholder(val):
            data[36] = val
        else:
            val = get_next_real_text(items, q24a_i)
            if val:
                data[36] = val

    # Q25: questionnaires familiarity
    q25i = fi("Is your site familiar with the planned questionnaires and scores")
    if q25i != -1:
        yn = get_yes_no_after(items, q25i)
        yn_idx = q25i
        for j in range(q25i + 1, min(q25i + 5, len(items))):
            if items[j]["checked"] is not None and "YES" in items[j]["text"].upper():
                yn_idx = j; break
        comment = get_comment_after_yn(items, yn_idx)
        data[37] = (f"{yn}. {comment}" if yn and comment else yn or comment)

    # Q26: treat acne without isotretinoin
    q26i = fi("Does your site treat subjects with the indication Moderate to Severe Acne Vulgaris who are not previously treated with Isotretinoin")
    data[38] = get_yes_no_after(items, q26i)

    # Q27: participated in acne studies
    q27i = fi("Has your site and Principal Investigator participated in previous Moderate to Severe Acne Vulgaris studies")
    data[39] = get_yes_no_after(items, q27i)

    # Q27 count – may be inline label OR a standalone paragraph right after YES/NO
    q27c_i = fi("If yes, please enter integer number here.")
    if q27c_i != -1:
        txt = items[q27c_i]["text"]
        val = get_inline_after(txt, "integer number here.")
        if val and not is_placeholder(val):
            data[40] = val
    # Fallback: standalone paragraph between Q27 YES/NO and Q27a
    if not data[40] and q27i != -1:
        q27a_search = fi("How many of these studies were Oral vs Topical")
        for j in range(q27i + 1, len(items)):
            if items[j]["checked"] is not None:
                continue  # skip YES/NO/Not sure checkboxes
            txt = items[j]["text"]
            if "Oral vs Topical" in txt or "How many of these studies" in txt.lower():
                break
            if txt and not is_placeholder(txt):
                data[40] = txt
                break

    # Q27a: oral vs topical (inline after "?")
    q27a_i = fi("How many of these studies were Oral vs Topical?")
    if q27a_i == -1:
        q27a_i = fi("a. How many of these studies were Oral vs Topical?")
    if q27a_i != -1:
        txt = items[q27a_i]["text"]
        val = get_inline_after(txt, "Oral vs Topical?")
        if val and not is_placeholder(val):
            # Strip any trailing placeholder label text appended after the real answer
            val = re.sub(r'\.\s*Please\b.*', '', val, flags=re.IGNORECASE).strip()
            if val:
                data[41] = val

    # ── Q28: competing studies ───────────────────────────────────────────────
    q28i = fi("Is your site currently participating or have plans to participate in any studies that may compete")
    data[42] = get_yes_no_after(items, q28i)

    q28c_i = fi("If yes, number of trials ongoing/planned:")
    if q28c_i != -1:
        txt = items[q28c_i]["text"]
        val = get_inline_after(txt, "ongoing/planned:")
        if val and not is_placeholder(val):
            data[43] = val

    # ── Q29: PCOS % ──────────────────────────────────────────────────────────
    q29i = fi("Approximately what percentage of subjects seen at your site have a diagnosis of PCOS")
    if q29i != -1:
        # Could be inline or next para
        txt = items[q29i]["text"]
        # Check if question and answer are in the same paragraph (rare)
        if "?" in txt:
            after_q = txt.split("?", 1)[-1].strip()
            if after_q and not is_placeholder(after_q):
                data[44] = after_q
        if not data[44]:
            val = get_next_real_text(items, q29i,
                stop_frags=["estimated number of adult subjects seen with",
                            "estimated number of pediatric subjects seen with"])
            if val:
                data[44] = val

    # ── Q30: adult subjects/month ─────────────────────────────────────────────
    q30i = fi("estimated number of adult subjects seen with Moderate to Severe Acne Vulgaris at your")
    if q30i != -1:
        txt = items[q30i]["text"]
        # Inline: "...per month: 80"
        val = get_inline_after(txt, "per month:")
        if val and not is_placeholder(val):
            data[45] = val
        else:
            val = get_next_real_text(items, q30i,
                stop_frags=["estimated number of pediatric subjects",
                            "Does your site have current subjects"])
            if val:
                data[45] = val

    # ── Q31: pediatric subjects/month ────────────────────────────────────────
    q31i = fi("estimated number of pediatric subjects (age 12-17) seen with Moderate to Severe Acne Vulgaris at your")
    if q31i != -1:
        txt = items[q31i]["text"]
        val = get_inline_after(txt, "per month:")
        if val and not is_placeholder(val):
            data[46] = val
        else:
            val = get_next_real_text(items, q31i,
                stop_frags=["Does your site have current subjects",
                            "your site's database"])
            if val:
                data[46] = val

    # ── Q32: current subjects in database ────────────────────────────────────
    q32i = fi("Does your site have current subjects with Moderate to Severe Acne Vulgaris in your site's database")
    data[47] = get_not_sure_after(items, q32i)

    # ── Q33: adult subjects in database ──────────────────────────────────────
    q33i = fi("estimated number of adult subjects with Moderate to Severe Acne Vulgaris in your site's database")
    if q33i != -1:
        txt = items[q33i]["text"]
        val = get_inline_after(txt, "database:")
        if val and not is_placeholder(val):
            data[48] = val
        else:
            val = get_next_real_text(items, q33i)
            if val:
                data[48] = val

    # ── Q34: pediatric subjects in database ───────────────────────────────────
    q34i = fi("estimated number of pediatric subjects (age 12-17) with Moderate to Severe Acne Vulgaris in your site's database")
    if q34i != -1:
        txt = items[q34i]["text"]
        val = get_inline_after(txt, "database:")
        if val and not is_placeholder(val):
            data[49] = val
        else:
            val = get_next_real_text(items, q34i)
            if val:
                data[49] = val

    # ── Q35: Asian % ─────────────────────────────────────────────────────────
    q35i = fi("What percentage of subjects with Acne Vulgaris seen at your")
    if q35i == -1:
        q35i = fi("Approximately what percentage of subjects with Acne Vulgaris")
    if q35i != -1:
        txt = items[q35i]["text"]
        val = get_inline_after(txt, "are Asian?")
        if not val:
            val = get_inline_after(txt, "Asian?")
        if val and not is_placeholder(val):
            data[50] = val
        else:
            val = get_next_real_text(items, q35i)
            if val:
                data[50] = val

    # ── Q36: expected randomizations ─────────────────────────────────────────
    q36_i = fi("How many subjects would your site expect to randomize in total")
    if q36_i == -1:
        q36_i = fi("Number of Pediatric Subjects (12-17)")

    # Q36a: pediatric
    q36a_i = fi("Number of Pediatric Subjects (12-17)")
    if q36a_i != -1:
        txt = items[q36a_i]["text"]
        val = get_inline_after(txt, "(12-17)")
        if val and not is_placeholder(val):
            data[51] = val.lstrip(":").strip()

    # Q36b: adult – search after Q36a to avoid matching Q30 text
    q36b_start = q36a_i if q36a_i != -1 else 0
    q36b_i = fi("Number of Adult Subjects ", q36b_start)
    if q36b_i == -1:
        q36b_i = fi("Number of Adult Subjects", q36b_start)
    if q36b_i != -1:
        txt = items[q36b_i]["text"]
        if "Number of Pediatric" not in txt:  # don't confuse with pediatric line
            val = get_inline_after(txt, "Number of Adult Subjects")
            if val and not is_placeholder(val):
                data[52] = val.lstrip(": ").strip()

    # Q36c: total
    q36c_i = fi("Total Number of Subjects")
    if q36c_i != -1:
        txt = items[q36c_i]["text"]
        val = get_inline_after(txt, "Total Number of Subjects")
        if val and not is_placeholder(val):
            val = val.lstrip(":").strip()
            # Strip spurious "Please " prefix from SDT concatenation artifact
            val = re.sub(r'^please\s+', '', val, flags=re.IGNORECASE).strip()
            if val and not is_placeholder(val):
                data[53] = val

    # ── Q37: recruitment support ─────────────────────────────────────────────
    q37i = fi("Would your site require patient recruitment support")
    data[54] = get_yes_no_after(items, q37i)

    # ── Q38: advertising methods ─────────────────────────────────────────────
    q38i = fi("What method of advertising would be most effective")
    if q38i != -1:
        data[55] = get_multi_checkbox(items, q38i,
                                       stop_frags=["screen failure", "please provide the screen fail"])

    # ── Q39: screen fail % ───────────────────────────────────────────────────
    q39i = fi("Please provide the screen fail rate percentage expectations")
    if q39i != -1:
        val = get_next_real_text(items, q39i)
        if val:
            data[56] = val

    # ── Q40: reason for screen fail ──────────────────────────────────────────
    q40i = fi("Based on inclusion and exclusion criteria, what would be the primary reason")
    if q40i != -1:
        val = get_next_real_text(items, q40i)
        if val:
            data[57] = val

    # ── Q41: discontinuation % ───────────────────────────────────────────────
    q41i = fi("Please provide your estimated subject discontinuation rate percentage")
    if q41i != -1:
        val = get_next_real_text(items, q41i)
        if val:
            data[58] = val

    # ── Q42: reasons for dropout ─────────────────────────────────────────────
    q42i = fi("Please list possible reasons for subjects discontinuing or dropping out")
    if q42i != -1:
        val = get_next_real_text(items, q42i)
        if val:
            data[59] = val

    # ── Q43: retention challenges ────────────────────────────────────────────
    q43i = fi("Does your site anticipate any challenges with subject retention or commitment due to the study duration")
    if q43i != -1:
        val = get_next_real_text(items, q43i)
        if val:
            data[60] = val

    # ── Q44: protocol challenges ─────────────────────────────────────────────
    q44i = fi("Please list any concerns or protocol challenges discovered during your site's review")
    if q44i != -1:
        val = get_next_real_text(items, q44i)
        if val:
            data[61] = val

    # Q44a: strategy
    q44a_i = fi("Briefly describe your site's plan or strategy to help overcome these challenges")
    if q44a_i != -1:
        val = get_next_real_text(items, q44a_i,
            stop_frags=["If your site treats pediatric subjects"])
        if val:
            data[62] = val

    # ── Q45: pediatric compliance challenges ─────────────────────────────────
    q45i = fi("If your site treats pediatric subjects, does your site anticipate any challenges")
    if q45i != -1:
        val = get_next_real_text(items, q45i,
            stop_frags=["Could your site enroll participants"])
        if val:
            data[63] = val

    # ── Q46: can enroll required lesion counts ───────────────────────────────
    q46i = fi("Could your site enroll participants with the required lesion counts")
    if q46i != -1:
        yn = get_yes_no_after(items, q46i)
        yn_idx = q46i
        for j in range(q46i + 1, min(q46i + 5, len(items))):
            if items[j]["checked"] is not None and "YES" in items[j]["text"].upper():
                yn_idx = j; break
        comment = get_comment_after_yn(items, yn_idx)
        data[64] = (f"{yn}. {comment}" if yn and comment else yn or comment)

    # ── Q47: eye exam capability ──────────────────────────────────────────────
    q47i = fi("Does your site have the capability to perform protocol")
    if q47i != -1:
        val = get_next_real_text(items, q47i,
            stop_frags=["qualified Optometrist or Ophthalmologist"])
        if val:
            data[65] = val

    # ── Q48: qualified optometrist ────────────────────────────────────────────
    q48i = fi("Does your site have a qualified Optometrist or Ophthalmologist on site")
    data[66] = get_yes_no_after(items, q48i)

    # Q48a: can contract (inline after "?")
    q48a_i = fi("If not, can your site contract an Optometrist or Ophthalmologist")
    if q48a_i != -1:
        txt = items[q48a_i]["text"]
        val = get_inline_after(txt, "Ophthalmologist?")
        if val and not is_placeholder(val):
            data[67] = val
        else:
            val = get_inline_after(txt, "Ophthalmologist")
            if val:
                val = val.lstrip("?").strip()
            if val and not is_placeholder(val):
                data[67] = val

    # Q48b: contract timeline (inline after "place?")
    q48b_i = fi("How long would this contract take to get in place")
    if q48b_i != -1:
        txt = items[q48b_i]["text"]
        val = get_inline_after(txt, "in place?")
        if not val:
            val = get_inline_after(txt, "in place")
            if val:
                val = val.lstrip("?").strip()
        if val and not is_placeholder(val):
            data[68] = val

    # ── Q49: recommend other dermatologists ──────────────────────────────────
    q49i = fi("Does your site recommend other Dermatologists or Dermatology Centers")
    if q49i != -1:
        lines = []
        for j in range(q49i + 1, len(items)):
            if items[j]["checked"] is not None:
                break
            txt = items[j]["text"]
            if "average contracting timeline" in _normalize(txt).lower():
                break
            if txt and not is_placeholder(txt):
                lines.append(txt)
        if lines:
            data[69] = "\n".join(lines)

    # ── Q50: contracting timeline ─────────────────────────────────────────────
    q50i = fi("What is your site's average contracting timeline for clinical studies")
    if q50i != -1:
        val = get_next_real_text(items, q50i)
        if val:
            # Strip spurious "Please " prefix from SDT concatenation artifact
            val = re.sub(r'^please\s+', '', val, flags=re.IGNORECASE).strip()
            if val and not is_placeholder(val):
                data[70] = val

    # ── Q51: per-subject investigator fee ─────────────────────────────────────
    q51i = fi("Is your site able to provide an estimate for a Per-Subject Investigator Fee")
    if q51i != -1:
        val = get_next_real_text(items, q51i)
        if val:
            data[71] = val

    # ── Q51a: other fees ─────────────────────────────────────────────────────
    q51a_i = fi("Any other fees")
    if q51a_i != -1:
        # Collect all lines until contracting contact question
        lines = []
        for j in range(q51a_i + 1, len(items)):
            if items[j]["checked"] is not None:
                break
            txt = items[j]["text"]
            if "Please enter contact information for responsible contracting person" in txt:
                break
            if txt and not is_placeholder(txt):
                lines.append(txt)
        if lines:
            data[72] = "\n".join(lines)

    # ── Q52: contracting contact ─────────────────────────────────────────────
    q52i = fi("Please enter contact information for responsible contracting person")
    if q52i != -1:
        fields = {}
        for j in range(q52i + 1, min(q52i + 10, len(items))):
            txt = items[j]["text"]
            if "Please select your site's availability" in txt:
                break
            for label in ["Last Name", "First Name", "Email Address", "Phone Number"]:
                if txt.startswith(label):
                    val = txt[len(label):].strip().lstrip(":").strip()
                    if val and not is_placeholder(val):
                        fields[label] = val
                    break
        lines = []
        for label in ["Last Name", "First Name", "Email Address", "Phone Number"]:
            if fields.get(label):
                lines.append(f"{label}: {fields[label]}")
        if lines:
            data[73] = "\n".join(lines)

    # ── IM availability ───────────────────────────────────────────────────────
    im_i = fi("Please select your site's availability for a 1-day Investigator Meeting")
    if im_i != -1:
        data[74] = get_multi_checkbox(items, im_i,
                                       stop_frags=["name and title of person completing"])

    # ── Name and title of person completing survey ────────────────────────────
    name_i = fi("Name and Title of person completing this survey")
    if name_i != -1:
        val = get_next_real_text(items, name_i)
        if val:
            data[75] = val

    # ── Date completed ────────────────────────────────────────────────────────
    date_i = fi("Date Questionnaire Completed")
    if date_i != -1:
        txt = items[date_i]["text"]
        val = get_inline_after(txt, "DD/MM/YY")
        if val and not is_placeholder(val):
            data[76] = val.lstrip(":").strip()
        else:
            val = get_next_real_text(items, date_i)
            if val:
                data[76] = val

    return data


# ── Write to Excel ───────────────────────────────────────────────────────────

def add_to_tracker(data, tracker_path):
    wb = openpyxl.load_workbook(tracker_path)
    ws = wb["All Responses"]

    # Check for duplicate: investigator name is in column A (data[0])
    investigator = data[0]
    if investigator:
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] and str(row[0]).strip().lower() == investigator.strip().lower():
                return None  # duplicate — do not write

    # Keep a timestamped backup of the original before writing
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = tracker_path.replace(".xlsx", f"_backup_{stamp}.xlsx")
    shutil.copy2(tracker_path, backup_path)

    # Find the last data row (first empty row after row 2)
    last_row = 2
    for row in ws.iter_rows(min_row=3):
        if any(cell.value is not None for cell in row):
            last_row = row[0].row
        else:
            break
    new_row = last_row + 1

    for col_idx, value in enumerate(data):
        if value is not None:
            ws.cell(row=new_row, column=col_idx + 1, value=value)

    wb.save(tracker_path)
    return new_row


# ── Logging helpers ──────────────────────────────────────────────────────────

COL_LABELS = [
    "Investigator", "Site Address", "State",
    "Q1 Feasible", "Q2 Interested", "Q2a Decline reason",
    "Q3 Board-certified", "Q4 Institution contact", "Q5 Site setting",
    "Q6 PI contact", "Q7 Staff", "Q8 Years conducting trials",
    "Q9 Acne studies", "Q9a Ph II count", "Q9b Ph III count",
    "Q10 GCP trained", "Q10 date", "Q11 Another investigator",
    "Q12 FDA audit", "Q12a 483",
    "Q13 Electronic files", "Q13a Which system", "Q13b Paper process",
    "Q14 CRA access", "Q15 Lockable storage", "Q16 Freezer",
    "Q17 Temp control", "Q18 Blood storage", "Q19 Blood to lab",
    "Q20 Acne photography", "Q20a House equipment",
    "Q21 Sebutape", "Q22 Sebumeter experience",
    "Q23 Calibrated sebumeter", "Q23a Require sebumeter",
    "Q24 Lesion count", "Q24a Who does lesion count",
    "Q25 Questionnaires", "Q26 Treat acne", "Q27 Participated acne studies",
    "Q27 count", "Q27a Oral/Topical",
    "Q28 Competing studies", "Q28 count",
    "Q29 PCOS %", "Q30 Adult/month", "Q31 Pediatric/month",
    "Q32 Current subjects in DB", "Q33 Adult in DB", "Q34 Pediatric in DB",
    "Q35 Asian %", "Q36a Pediatric", "Q36b Adult", "Q36c Total",
    "Q37 Recruitment support", "Q38 Advertising",
    "Q39 Screen fail %", "Q40 Reason screen fail",
    "Q41 Discontinuation %", "Q42 Reasons dropout",
    "Q43 Retention challenges", "Q44 Protocol challenges", "Q44a Strategy",
    "Q45 Pediatric challenges", "Q46 Lesion counts enrollable",
    "Q47 Eye exam", "Q48 Optometrist", "Q48a Can contract",
    "Q48b Contract timeline", "Q49 Recommend others",
    "Q50 Contracting timeline", "Q51 Per-subject fee", "Q51a Other fees",
    "Q52 Contracting contact", "IM availability",
    "Name/title completing", "Date completed", "Site comments",
]

def print_data(data):
    print("─" * 70)
    for i, (label, value) in enumerate(zip(COL_LABELS, data)):
        if value is not None:
            short_val = str(value)[:120].replace("\n", " | ")
            print(f"  Col {i:2d} {label}: {short_val}")
    print("─" * 70)


# ── Single-file processor ─────────────────────────────────────────────────────

def process_file(docx_path, dry_run, progress_prefix=""):
    """Parse one docx and optionally write to the tracker.

    Returns a dict with keys: fname, status ('ok'/'skipped'/'error'), message.
    """
    fname = os.path.basename(docx_path)
    last_name = fname.split()[0]

    print(f"\n{progress_prefix}Processing: {fname}")
    print(f"  Looking up '{last_name}' in master spreadsheet…")

    master = find_master_info(last_name)
    if not master:
        msg = f"'{last_name}' not found in master spreadsheet — skipping."
        print(f"  WARNING: {msg}")
        return {"fname": fname, "status": "skipped", "message": msg}

    print(f"  Found: {master['investigator']}  |  {master['state']}")
    print(f"  Parsing questionnaire…")

    try:
        data = parse_docx(docx_path)
    except Exception as exc:
        msg = f"Parse error: {exc}"
        print(f"  ERROR: {msg}")
        return {"fname": fname, "status": "error", "message": msg}

    data[0] = master["investigator"]
    data[1] = master["site_address"]
    data[2] = master["state"]

    print_data(data)

    if dry_run:
        print("  [DRY RUN] Not written to Excel.")
        return {"fname": fname, "status": "ok", "message": "dry run"}

    try:
        new_row = add_to_tracker(data, TRACKER_FILE)
    except Exception as exc:
        msg = f"Excel write error: {exc}"
        print(f"  ERROR: {msg}")
        return {"fname": fname, "status": "error", "message": msg}

    if new_row is None:
        msg = "Already exists in tracker — skipped (duplicate)."
        print(f"  SKIPPED: {msg}")
        return {"fname": fname, "status": "skipped", "message": msg}

    print(f"  ✓ Written to row {new_row} of '{os.path.basename(TRACKER_FILE)}'.")
    return {"fname": fname, "status": "ok", "message": f"row {new_row}"}


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "path",
        help="Path to a single .docx file OR a directory containing .docx files",
    )
    parser.add_argument(
        "--dry-run", "--preview",
        action="store_true",
        help="Print extracted data without writing to Excel",
    )
    args = parser.parse_args()

    # Resolve path
    target = args.path
    if not os.path.exists(target):
        target = os.path.join(BASE_DIR, args.path)
    if not os.path.exists(target):
        sys.exit(f"ERROR: path not found: {args.path}")

    # Collect docx files to process
    if os.path.isdir(target):
        docx_files = sorted(
            os.path.join(target, f)
            for f in os.listdir(target)
            if f.lower().endswith(".docx") and not f.startswith("~$")
        )
        if not docx_files:
            sys.exit(f"ERROR: no .docx files found in {target}")
    else:
        docx_files = [target]

    total = len(docx_files)
    width = len(str(total))   # digit width for zero-padding

    if args.dry_run:
        print(f"[DRY RUN MODE] No data will be written to Excel.\n")
    print(f"Found {total} questionnaire(s) to process.\n")

    results = []
    for idx, docx_path in enumerate(docx_files, start=1):
        prefix = f"({idx:{width}d}/{total}) "
        result = process_file(docx_path, dry_run=args.dry_run, progress_prefix=prefix)
        results.append(result)

    # ── Summary ──────────────────────────────────────────────────────────────
    ok       = [r for r in results if r["status"] == "ok"]
    skipped  = [r for r in results if r["status"] == "skipped"]
    errors   = [r for r in results if r["status"] == "error"]

    print("\n" + "═" * 70)
    print(f"SUMMARY  ({total} file(s) processed)")
    print("═" * 70)
    print(f"  ✓ Succeeded : {len(ok)}")
    if skipped:
        print(f"  ⚠ Skipped   : {len(skipped)}")
        for r in skipped:
            print(f"      {r['fname']} — {r['message']}")
    if errors:
        print(f"  ✗ Errors    : {len(errors)}")
        for r in errors:
            print(f"      {r['fname']} — {r['message']}")
    if not args.dry_run and ok:
        print(f"\nPlease open '{os.path.basename(TRACKER_FILE)}' to review and add any manual corrections.")
    print("═" * 70)


if __name__ == "__main__":
    main()
